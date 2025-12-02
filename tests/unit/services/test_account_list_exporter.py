"""Unit tests for the account list artifact exporter."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from i4g.services.account_list.exporters import AccountListExporter
from i4g.services.account_list.models import AccountListResult, FinancialIndicator


def _build_result() -> AccountListResult:
    return AccountListResult(
        request_id="acc-export-1",
        generated_at=datetime(2025, 11, 25, 10, 30, tzinfo=timezone.utc),
        indicators=[
            FinancialIndicator(
                category="bank",
                item="Example Bank",
                type="bank_account",
                number="****1111",
                source_case_id="case-1",
                metadata={"note": "primary"},
            )
        ],
        sources=[],
        warnings=[],
        metadata={"indicator_count": 1},
    )


def test_exporter_writes_csv(tmp_path):
    exporter = AccountListExporter(base_dir=tmp_path)
    result, warnings = exporter.export(_build_result(), ["csv"])

    assert not warnings
    assert "csv" in result
    csv_path = Path(result["csv"])
    assert csv_path.exists()
    content = csv_path.read_text(encoding="utf-8")
    assert "Example Bank" in content


def test_exporter_writes_xlsx(tmp_path):
    exporter = AccountListExporter(base_dir=tmp_path)
    paths, warnings = exporter.export(_build_result(), ["xlsx"])

    assert not warnings
    xlsx_path = Path(paths["xlsx"])
    assert xlsx_path.exists()
    workbook = load_workbook(filename=xlsx_path)
    sheet = workbook["Indicators"]
    assert sheet.max_row == 2
    assert sheet["C2"].value == "Example Bank"


def test_exporter_writes_pdf(tmp_path):
    exporter = AccountListExporter(base_dir=tmp_path)
    paths, warnings = exporter.export(_build_result(), ["pdf"])

    assert not warnings
    pdf_path = Path(paths["pdf"])
    assert pdf_path.exists()
    assert pdf_path.read_bytes().startswith(b"%PDF")


def test_exporter_handles_unknown_format(tmp_path):
    exporter = AccountListExporter(base_dir=tmp_path)
    result, warnings = exporter.export(_build_result(), ["svg"])
    assert result == {}
    assert warnings


def test_exporter_prefers_drive_link(monkeypatch, tmp_path):
    exporter = AccountListExporter(base_dir=tmp_path)
    exporter._drive_folder_id = "folder-id"
    exporter._drive_service = object()

    called = {}

    def _mock_upload(self, path, content_type):  # noqa: ANN001 - test helper
        called["hit"] = True
        return "https://drive.example/file"

    monkeypatch.setattr(AccountListExporter, "_upload_to_drive", _mock_upload)
    assert AccountListExporter._upload_to_drive is _mock_upload

    result, warnings = exporter.export(_build_result(), ["json"])
    assert called.get("hit") is True
    assert not warnings
    assert result["json"] == "https://drive.example/file"


def test_exporter_records_drive_warning(monkeypatch, tmp_path):
    exporter = AccountListExporter(base_dir=tmp_path)
    exporter._drive_folder_id = "folder-id"
    exporter._drive_service = object()

    def _fail_upload(self, *_args, **_kwargs):  # noqa: ANN001 - stub
        raise RuntimeError("drive boom")

    monkeypatch.setattr(AccountListExporter, "_upload_to_drive", _fail_upload)

    artifacts, warnings = exporter.export(_build_result(), ["json"])

    assert artifacts["json"].endswith(".json")
    assert any("Drive upload failed" in warning for warning in warnings)


def test_exporter_records_gcs_warning(tmp_path):
    exporter = AccountListExporter(base_dir=tmp_path)

    class _FailBlob:
        def upload_from_file(self, *_args, **_kwargs):  # noqa: ANN001
            raise RuntimeError("gcs boom")

    class _FailBucket:
        def blob(self, *_args, **_kwargs):  # noqa: ANN001
            return _FailBlob()

    exporter._reports_bucket = "test-bucket"
    exporter._bucket = _FailBucket()

    artifacts, warnings = exporter.export(_build_result(), ["csv"])

    assert artifacts["csv"].endswith(".csv")
    assert any("GCS upload failed" in warning for warning in warnings)
