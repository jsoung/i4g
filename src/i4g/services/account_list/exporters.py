"""Artifact exporters for account list extraction results."""

from __future__ import annotations

import csv
import json
import logging
import mimetypes
from datetime import timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from i4g.settings import Settings, get_settings

from .models import AccountListResult

LOGGER = logging.getLogger(__name__)

try:  # pragma: no cover - optional dependency
    from google.cloud import storage
except ImportError:  # pragma: no cover - local/dev environments may not install GCS
    storage = None


class AccountListExporter:
    """Generate artifact files (CSV, JSON, etc.) for extraction results."""

    _DRIVE_SCOPES = (
        "https://www.googleapis.com/auth/drive.file",
        "https://www.googleapis.com/auth/cloud-platform",
    )

    def __init__(
        self,
        *,
        settings: Settings | None = None,
        base_dir: Path | None = None,
        drive_service: Any | None = None,
    ) -> None:
        self.settings = settings or get_settings()
        # Determine reports bucket early so we can choose a writable local
        # directory when uploads will be sent to GCS. Using a temporary
        # directory prevents attempts to create folders under the package
        # install path (for example `/usr/local/lib/data`) which is often
        # not writable in container runtimes.
        storage_settings = self.settings.storage
        self._reports_bucket = storage_settings.reports_bucket

        if base_dir is not None:
            self.base_dir = base_dir
        elif self._reports_bucket:
            # When a remote reports bucket is configured, write artifacts to
            # a writable temporary directory and upload them to GCS.
            import tempfile

            self.base_dir = Path(tempfile.mkdtemp(prefix="i4g_reports_"))
        else:
            self.base_dir = self.settings.data_dir / "reports" / "account_list"

        self.base_dir.mkdir(parents=True, exist_ok=True)
        self._artifact_prefix = self.settings.account_list.artifact_prefix or "account_list"
        self._drive_folder_id = self.settings.account_list.drive_folder_id
        if self._reports_bucket:
            if storage is None:
                raise RuntimeError("google-cloud-storage required for reports bucket uploads")
            project = storage_settings.firestore_project
            client = storage.Client(project=project)  # type: ignore[attr-defined]
            self._bucket = client.bucket(self._reports_bucket)
        else:
            self._bucket = None
        self._content_types = {
            "csv": "text/csv",
            "json": "application/json",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "pdf": "application/pdf",
        }
        self._drive_service = drive_service or self._build_drive_client()

    def export(self, result: AccountListResult, formats: Iterable[str]) -> tuple[Dict[str, str], List[str]]:
        """Write artifacts for the requested formats and return their paths/warnings."""

        artifacts: Dict[str, str] = {}
        warnings: List[str] = []
        for fmt in formats:
            normalized = fmt.lower().strip()
            if not normalized:
                continue
            handler = getattr(self, f"_export_{normalized}", None)
            if handler is None:
                LOGGER.warning("Unsupported account list artifact format: %s", normalized)
                warnings.append(f"Unsupported artifact format skipped: {normalized}")
                continue
            try:
                path = handler(result)
                artifacts[normalized] = self._finalize_artifact(path, warnings, format_name=normalized)
            except Exception as exc:  # pragma: no cover - file I/O edge cases
                LOGGER.exception("Failed to generate %s artifact", normalized)
                warnings.append(f"{normalized.upper()} artifact generation failed: {exc}")
        return artifacts, warnings

    # ------------------------------------------------------------------
    # Individual format handlers
    # ------------------------------------------------------------------

    def _export_csv(self, result: AccountListResult) -> Path:
        """Write indicators to a CSV file and return the path."""

        path = self._build_path(result, suffix="csv")
        fieldnames = [
            "category",
            "type",
            "item",
            "number",
            "source_case_id",
            "metadata",
        ]
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            for indicator in result.indicators:
                payload = indicator.model_dump()
                metadata = payload.get("metadata") or {}
                writer.writerow(
                    {
                        "category": payload.get("category"),
                        "type": payload.get("type"),
                        "item": payload.get("item"),
                        "number": payload.get("number"),
                        "source_case_id": payload.get("source_case_id"),
                        "metadata": json.dumps(metadata, ensure_ascii=False),
                    }
                )
        return path

    def _export_json(self, result: AccountListResult) -> Path:
        """Write the entire result payload to a JSON file for downstream use."""

        path = self._build_path(result, suffix="json")
        with path.open("w", encoding="utf-8") as handle:
            json.dump(result.model_dump(mode="json"), handle, ensure_ascii=False, indent=2)
        return path

    def _export_xlsx(self, result: AccountListResult) -> Path:
        """Write indicators to an XLSX workbook."""

        path = self._build_path(result, suffix="xlsx")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Indicators"
        headers = ["Category", "Type", "Item", "Number", "Source Case ID"]
        sheet.append(headers)
        header_font = Font(bold=True)
        for idx in range(1, len(headers) + 1):
            sheet.cell(row=1, column=idx).font = header_font
        for indicator in result.indicators:
            sheet.append(
                [
                    indicator.category,
                    indicator.type,
                    indicator.item,
                    indicator.number,
                    indicator.source_case_id or "",
                ]
            )
        summary = workbook.create_sheet("Summary")
        summary["A1"] = "Request ID"
        summary["B1"] = result.request_id
        summary["A2"] = "Generated At"
        summary["B2"] = result.generated_at.astimezone(timezone.utc).isoformat()
        summary["A3"] = "Indicator Count"
        summary["B3"] = len(result.indicators)
        workbook.save(path)
        return path

    def _export_pdf(self, result: AccountListResult) -> Path:
        """Render indicators into a simple PDF table."""

        path = self._build_path(result, suffix="pdf")
        doc = SimpleDocTemplate(str(path), pagesize=LETTER)
        styles = getSampleStyleSheet()
        story = [Paragraph("Account List Extraction", styles["Title"]), Spacer(1, 12)]
        summary_text = f"Run: {result.request_id} — {result.generated_at.astimezone(timezone.utc).isoformat()}"
        story.append(Paragraph(summary_text, styles["Normal"]))
        story.append(Spacer(1, 12))
        table_data = [["Category", "Type", "Item", "Number", "Source"]]
        for indicator in result.indicators:
            table_data.append(
                [
                    indicator.category,
                    indicator.type,
                    indicator.item,
                    indicator.number,
                    indicator.source_case_id or "",
                ]
            )
        if len(table_data) == 1:
            table_data.append(["—", "—", "No indicators", "—", "—"])
        table = Table(table_data, hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ]
            )
        )
        story.append(table)
        doc.build(story)
        return path

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _build_path(self, result: AccountListResult, *, suffix: str) -> Path:
        timestamp = result.generated_at.astimezone(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        safe_request_id = result.request_id.replace("/", "-")
        filename = f"{safe_request_id}_{timestamp}.{suffix}"
        return self.base_dir / filename

    def _finalize_artifact(self, local_path: Path, warnings: List[str], *, format_name: str) -> str:
        content_type = self._content_types.get(local_path.suffix.lstrip(".").lower())
        content_type = content_type or mimetypes.guess_type(local_path.name)[0] or "application/octet-stream"

        if self._drive_folder_id and self._drive_service:
            try:
                return self._upload_to_drive(local_path, content_type)
            except Exception as exc:  # pragma: no cover - Drive client failures
                warning = f"Drive upload failed for {format_name} artifact: {exc}"
                warnings.append(warning)
                LOGGER.warning(warning)

        if self._bucket is not None:
            try:
                return self._upload_to_gcs(local_path, content_type)
            except Exception as exc:  # pragma: no cover - GCS failures
                warning = f"GCS upload failed for {format_name} artifact: {exc}"
                warnings.append(warning)
                LOGGER.warning(warning)

        return str(local_path)

    def _upload_to_gcs(self, local_path: Path, content_type: str) -> str:
        blob_path = f"{self._artifact_prefix}/{local_path.name}"
        blob = self._bucket.blob(blob_path)  # type: ignore[union-attr]
        with local_path.open("rb") as handle:
            blob.upload_from_file(handle, rewind=True, content_type=content_type)
        return f"gs://{self._reports_bucket}/{blob_path}"

    def _build_drive_client(self):
        if not self._drive_folder_id:
            return None
        try:
            import google.auth
            from googleapiclient.discovery import build
        except ImportError:  # pragma: no cover - optional dependency missing
            LOGGER.warning("Drive client unavailable: google-auth or googleapiclient missing")
            return None
        try:
            credentials, _ = google.auth.default(scopes=list(self._DRIVE_SCOPES))
        except Exception:  # pragma: no cover - ADC failure surfaces in logs
            LOGGER.exception("Unable to load ADC credentials for Drive uploads")
            return None
        try:
            return build("drive", "v3", credentials=credentials, cache_discovery=False)
        except Exception:  # pragma: no cover - Drive client init issues
            LOGGER.exception("Failed to build Drive client")
            return None

    def _upload_to_drive(self, local_path: Path, content_type: str) -> str:
        if not self._drive_folder_id or not self._drive_service:
            raise RuntimeError("Drive uploads not configured")
        try:
            from googleapiclient.http import MediaFileUpload
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("googleapiclient missing for Drive uploads") from exc

        media = MediaFileUpload(str(local_path), mimetype=content_type, resumable=False)
        metadata = {
            "name": local_path.name,
            "parents": [self._drive_folder_id],
        }
        request = self._drive_service.files().create(  # type: ignore[union-attr]
            body=metadata,
            media_body=media,
            fields="id,webViewLink,webContentLink",
            supportsAllDrives=True,
        )
        file_info = request.execute()
        link = file_info.get("webViewLink") or file_info.get("webContentLink")
        if not link and file_info.get("id"):
            link = f"https://drive.google.com/file/d/{file_info['id']}/view"
        if not link:
            raise RuntimeError("Drive upload returned no shareable link")
        return link


__all__ = ["AccountListExporter"]
